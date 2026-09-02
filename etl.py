# -*- coding: utf-8 -*-
"""
ETL - Projecao de Deficit GDF 2026 (arquivo unico).
Conecta no Oracle, roda as 2 consultas do Discoverer, monta os dados e gera a
planilha Analise_Deficit_2026_AAAA-MM-DD.xlsx com os Comparativos 1, 2 e 3.

Uso (Prompt de Comando, na pasta do projeto):
    set ORACLE_PWD=sua_senha
    py etl.py

Arquivos que devem estar na mesma pasta: etl.py, "PROJECAO DESPESAS.sql", "LIQUIDADO.sql".
A senha vem da variavel ORACLE_PWD e NAO fica salva em arquivo.
"""
import os
import sys
import json
import re
from datetime import date
import oracledb
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ======================= CONFIGURACAO =======================
# Dados de conexao ficam em config.py (que NAO vai para o GitHub).
try:
    from config import INSTANT_CLIENT, HOST, PORT, SID, USER
except ImportError:
    sys.exit("Falta o arquivo config.py. Copie o config_exemplo.py para config.py "
             "e preencha com os dados de conexao (host, porta, SID, usuario, Instant Client).")

SQL_QDD_FILE = "PROJECAO DESPESAS.sql"
SQL_LIQ_FILE = "LIQUIDADO.sql"

# Saida com a data do dia (um arquivo novo a cada rodada). Para dia/mes: "%d-%m".
SAIDA = f"Analise_Deficit_2026_{date.today():%Y-%m-%d}.xlsx"
# ================================================================================

QDD_COLS = ['COEXERCICIO', 'COFONTEMAE', 'CONATUREZA', 'NOSUBTITULO', 'NOUO', 'COUO',
            'PT', 'IDUSO', 'DOTMAISALT', 'DESPAUT', 'ALT', 'LEI', 'CONT105',
            'LIQ', 'EMP', 'DISP', 'CONT104']
LIQ_COLS = ['COEXERCICIO', 'COFONTEMAE', 'INMES', 'PT', 'LIQ']

# =============================== RESHAPE (long -> arquivos-fonte) ==============================
ANO_LABELS = ['LEI', 'ALTERAÇÃO', 'COTA', 'CONTINGENCIADO', 'DESP AUT.',
              'EMPENHADO', 'LIQUIDADO', 'DISPONÍVEL', 'Dotação Autorizada SUM']

def montar_qdd(dfq):
    """dfq: colunas COEXERCICIO, COFONTEMAE, CONATUREZA, NOSUBTITULO, NOUO, COUO,
    PT, IDUSO, DOTMAISALT, DESPAUT, ALT, LEI, CONT105, LIQ, EMP, DISP, CONT104"""
    key = ['COUO', 'NOUO', 'PT', 'NOSUBTITULO', 'CONATUREZA', 'COFONTEMAE', 'IDUSO']
    def bloco(r):
        # 9 valores na ordem do .xls
        return [r['LEI'], r['ALT'], r['CONT105'], r['CONT104'], r['DESPAUT'],
                r['EMP'], r['LIQ'], r['DISP'], r['DOTMAISALT']]
    linhas = []
    # cabecalhos (2 linhas)
    h0 = [None]*7 + ['2025']*9 + ['2026']*9
    h1 = [None]*7 + ANO_LABELS + ANO_LABELS
    linhas.append(h0); linhas.append(h1)
    # agrupa por chave; separa 2025 e 2026
    dfq = dfq.copy()
    for kcol in key:
        pass
    grp = {}
    for _, r in dfq.iterrows():
        k = tuple(r[c] for c in key)
        grp.setdefault(k, {})[int(r['COEXERCICIO'])] = r
    zero9 = [0]*9
    for k, anos in grp.items():
        b25 = bloco(anos[2025]) if 2025 in anos else zero9
        b26 = bloco(anos[2026]) if 2026 in anos else zero9
        linhas.append(list(k) + b25 + b26)
    return linhas

def montar_liquidado(dfl):
    """dfl: colunas COFONTEMAE, INMES, PT, LIQ"""
    # posicoes fixas: col2=mes1 ... col13=mes12
    linhas = []
    h0 = [None, None] + ['2026']*12
    h1 = [None, None] + ['LIQUIDADO']*12
    h2 = [None, None] + [str(m) for m in range(1, 13)]
    linhas.append(h0); linhas.append(h1); linhas.append(h2)
    tab = {}
    for _, r in dfl.iterrows():
        chave = (r['PT'], r['COFONTEMAE'])
        mes = int(r['INMES'])
        tab.setdefault(chave, {})[mes] = r['LIQ']
    for (pt, fonte), meses in tab.items():
        linha = [pt, fonte] + [meses.get(m, 0) for m in range(1, 13)]
        linhas.append(linha)
    return linhas

def salvar_aba(linhas, caminho, nome_aba):
    wb = Workbook(); ws = wb.active; ws.title = nome_aba
    for lin in linhas:
        ws.append(lin)
    wb.save(caminho)

# =============================== GERACAO DA ANALISE (ex-build_deficit) =========================
def gerar_analise(SRC, SRC_LIQ, OUT):
    QDD_NAME = os.path.basename(SRC)
    LIQ_NAME = os.path.basename(SRC_LIQ)
    # ================= PREMISSA DE CORTE TEMPORAL (edite 1 número por mês) =================
    # MESES_FECHADOS = quantos meses de 2026 já estão fechados no LIQUIDADO.
    #   jul fechado -> 7 | ago fechado -> 8 | set fechado -> 9 ...
    # Tudo abaixo (janela média/mediana, projeção, rótulos, gráficos) deriva deste número.
    MESES_FECHADOS = 8
    # ---------- derivados (não editar) ----------
    _MN = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    NCLOSED = MESES_FECHADOS
    WIN_START = 3                        # abril (índice 0-based): início fixo da janela média/mediana
    WIN = NCLOSED - WIN_START            # nº de meses na janela (abr .. último fechado)
    REST = 12 - NCLOSED                  # meses restantes a projetar
    FAT = 12 / NCLOSED                   # fator de anualização do empenho
    LAST_M = _MN[NCLOSED - 1]            # último mês fechado, ex.: 'ago'
    ACUM_LBL = f'jan–{LAST_M}'      # 'jan–ago'
    WIN_LBL = f'abr–{LAST_M}'       # 'abr–ago'
    REST_LBL = f'{_MN[NCLOSED]}–dez' if NCLOSED < 12 else 'sem meses restantes'
    PCT_DEC = f'{NCLOSED / 12 * 100:.1f}'.replace('.', ',')  # '66,7'

    TARGETS = ['15.452.6209.2079.6118', '26.453.6216.4202.0002', '26.453.6216.2455.0002',
               '10.302.6202.4206.0001', '10.302.6202.4206.0002',
               '12.361.6221.4202.0004', '12.453.6221.4202.0010',
               '08.244.6228.4272.0001', '08.306.6228.4175.0002',
               '08.245.6228.9073.0003', '08.245.6228.9071.0003',
               '08.244.6228.4162.0008', '12.122.6221.4091.0096',
               '26.453.6216.2455.0001', '26.782.6216.2885.0001',
               '26.782.6216.4195.0001', '26.453.6216.2756.6137',
               '15.452.6209.8508.0002', '17.512.6209.2903.0001',
               '06.422.6217.2540.0002', '06.421.6217.2727.0006',
               '04.122.8203.2990.0006', '04.122.0001.9126.0001',
               '04.122.8203.2984.0001', '04.122.8203.2422.0006',
               '04.122.8203.2990.0008']
    SHEETN = {'15.452.6209.2079.6118': 'PT_2079_6118',
              '26.453.6216.4202.0002': 'PT_4202_0002',
              '26.453.6216.2455.0002': 'PT_2455_0002',
              '10.302.6202.4206.0001': 'PT_4206_0001',
              '10.302.6202.4206.0002': 'PT_4206_0002',
              '12.361.6221.4202.0004': 'PT_4202_0004',
              '12.453.6221.4202.0010': 'PT_4202_0010',
              '08.244.6228.4272.0001': 'PT_4272_0001',
              '08.306.6228.4175.0002': 'PT_4175_0002',
              '08.245.6228.9073.0003': 'PT_9073_0003',
              '08.245.6228.9071.0003': 'PT_9071_0003',
              '08.244.6228.4162.0008': 'PT_4162_0008',
              '12.122.6221.4091.0096': 'PT_4091_0096',
              '26.453.6216.2455.0001': 'PT_2455_0001',
              '26.782.6216.2885.0001': 'PT_2885_0001',
              '26.782.6216.4195.0001': 'PT_4195_0001',
              '26.453.6216.2756.6137': 'PT_2756_6137',
              '15.452.6209.8508.0002': 'PT_8508_0002',
              '17.512.6209.2903.0001': 'PT_2903_0001',
              '06.422.6217.2540.0002': 'PT_2540_0002',
              '06.421.6217.2727.0006': 'PT_2727_0006',
              '04.122.8203.2990.0006': 'PT_2990_0006',
              '04.122.0001.9126.0001': 'PT_9126_0001',
              '04.122.8203.2984.0001': 'PT_2984_0001',
              '04.122.8203.2422.0006': 'PT_2422_0006',
              '04.122.8203.2990.0008': 'PT_2990_0008'}
    DESCS = {'15.452.6209.2079.6118': 'SLU — Manutenção das Atividades de Limpeza Pública',
             '26.453.6216.4202.0002': 'SEMOB — Concessão de Passe Livre PNE',
             '26.453.6216.2455.0002': 'SEMOB — Equilíbrio Financeiro do STPC',
             '10.302.6202.4206.0001': 'FSDF — Contratos de Gestão IGES-DF',
             '10.302.6202.4206.0002': 'FSDF — Contratos de Gestão HCB',
             '12.361.6221.4202.0004': 'SEEDF — Passe Livre Estudantil (Ensino Fundamental)',
             '12.453.6221.4202.0010': 'SEEDF — Passe Livre Estudantil (PLE Outros)',
             '08.244.6228.4272.0001': 'SEDES — Cartão Prato Cheio',
             '08.306.6228.4175.0002': 'SEDES — Restaurantes Comunitários',
             '08.245.6228.9073.0003': 'FAS — Bloco Proteção Social Especial',
             '08.245.6228.9071.0003': 'FAS — Bloco Proteção Social Básica',
             '08.244.6228.4162.0008': 'FCEP — Transferência de Renda (DF Social)',
             '12.122.6221.4091.0096': 'SEEDF — Educador Social Voluntário (ESV)',
             '26.453.6216.2455.0001': 'SEMOB — Equilíbrio Financeiro do STPC (0001)',
             '26.782.6216.2885.0001': 'DER — Manutenção de Máquinas e Equipamentos',
             '26.782.6216.4195.0001': 'DER — Conservação de Rodovias',
             '26.453.6216.2756.6137': 'Metrô-DF — Manutenção do Sistema Ferroviário',
             '15.452.6209.8508.0002': 'NOVACAP — Manutenção de Áreas Urbanizadas e Ajardinadas',
             '17.512.6209.2903.0001': 'NOVACAP — Manutenção de Redes de Águas Pluviais',
             '06.422.6217.2540.0002': 'SEAPE — Fornecimento de Alimentação aos Presidiários',
             '06.421.6217.2727.0006': 'SEAPE — Manutenção do Sistema Penitenciário',
             '04.122.8203.2990.0006': 'SEEC — Manutenção de Bens Imóveis do GDF (Vigilância)',
             '04.122.0001.9126.0001': 'SEEC — Aporte da Contribuição Mensal do GDF',
             '04.122.8203.2984.0001': 'SEEC — Manutenção da Frota Oficial de Veículos',
             '04.122.8203.2422.0006': 'SEEC — Concessão de Bolsa Estágio',
             '04.122.8203.2990.0008': 'SEEC — Manutenção de Bens Imóveis do GDF (Limpeza)'}

    # Ordena os PTs pela descrição (col B do Consolidado) — reordena abas e linhas
    TARGETS = sorted(TARGETS, key=lambda pt: DESCS[pt].lower())

    # ---------- Ler QDD (.xls) via pandas ----------
    # PTs transformados durante 2025 → consolidados sob o sucessor (só execução 2025)
    REMAP = {'08.306.6228.4173.0003': '08.244.6228.4272.0001'}
    dfq = pd.ExcelFile(SRC).parse('2025 e 2026', header=None)
    # metadados 0-6; 2025: 7-15 (LEI,ALT,COTA,CONT,DAUT,EMP,LIQ,DISP,DotAutSUM); 2026: 16-24
    header1 = list(dfq.iloc[0])
    header2 = list(dfq.iloc[1])
    rows_data = []
    fontes = {pt: set() for pt in TARGETS}
    for _, r in dfq.iterrows():
        pt_raw = r[2]
        pt = REMAP.get(pt_raw, pt_raw)
        if pt in TARGETS:
            vals = []
            for c in range(25):
                v = r[c]
                if c == 2:
                    vals.append(pt)  # código consolidado (sucessor)
                elif pd.isna(v):
                    vals.append(None)
                elif c in (0, 3):
                    vals.append(str(v))
                elif c in (4, 5, 6):
                    vals.append(str(int(v)) if isinstance(v, float) and float(v).is_integer() else str(v))
                elif c == 1:
                    vals.append(str(v))
                else:
                    vals.append(float(v))
            rows_data.append(vals)
            fontes[pt].add(str(int(r[5]))[:3])
    fontes = {pt: sorted(v) for pt, v in fontes.items()}

    # ---------- Definição das ABAS (consolidando a Ação 4202 — Passe Livre) ----------
    def action_of(code):
        p = str(code).split('.')
        return p[3] if len(p) > 3 else ''

    # Ações agregadas numa única aba/linha: (sheet, descrição, rótulo do consolidado)
    GROUP_ACTIONS = {
        '4202': ('PT_4202_Consolidado', 'Passe Livre — Consolidado (Ação 4202)', 'Ação 4202 (Passe Livre)'),
        '2455': ('PT_2455_Consolidado', 'STPC — Equilíbrio Financeiro Consolidado (Ação 2455)', 'Ação 2455 (Equilíbrio STPC)'),
    }
    TABS = []
    for pt in TARGETS:
        if action_of(pt) in GROUP_ACTIONS:
            continue
        TABS.append({'key': pt, 'match': pt, 'desc': DESCS[pt], 'sheet': SHEETN[pt],
                     'fontes': fontes[pt], 'group': False, 'action': action_of(pt)})
    for act, (sheet, desc, key) in GROUP_ACTIONS.items():
        members = sorted([pt for pt in TARGETS if action_of(pt) == act])
        if not members:
            continue
        gf = sorted(set().union(*[set(fontes[pt]) for pt in members]))
        TABS.append({'key': key, 'match': f'*.{act}.*', 'desc': desc, 'sheet': sheet,
                     'fontes': gf, 'group': True, 'members': members, 'action': act})
    TABS.sort(key=lambda t: t['desc'].lower())

    wb = openpyxl.Workbook()
    F = 'Arial'
    def f(sz=10, b=False, color='000000'): return Font(name=F, size=sz, bold=b, color=color)
    NUM = '#,##0.00;[Red](#,##0.00)'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill('solid', fgColor='1F4E78')
    h2fill = PatternFill('solid', fgColor='D9E1F2')
    yfill = PatternFill('solid', fgColor='FFFF00')
    gfill = PatternFill('solid', fgColor='F2F2F2')

    # ---------- Aba 1: QDD_Filtrado ----------
    d = wb.active
    d.title = 'QDD_Filtrado'
    for c, v in enumerate(header1, 1):
        d.cell(1, c, None if (isinstance(v, float) and pd.isna(v)) else v)
    for c, v in enumerate(header2, 1):
        d.cell(2, c, None if (isinstance(v, float) and pd.isna(v)) else v)
    d.merge_cells('H1:P1'); d.merge_cells('Q1:Y1')
    d['H1'] = '2025'; d['Q1'] = '2026'
    for c in range(1, 26):
        for r in (1, 2):
            cell = d.cell(r, c)
            cell.font = f(10, True, 'FFFFFF'); cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    for ri, row in enumerate(rows_data, 3):
        for c, v in enumerate(row, 1):
            cell = d.cell(ri, c, v)
            cell.font = f(9); cell.border = border
            if c >= 8: cell.number_format = NUM
    widths = [8, 30, 22, 45, 9, 12, 6] + [14]*18
    for c, w in enumerate(widths, 1): d.column_dimensions[get_column_letter(c)].width = w
    d.freeze_panes = 'A3'

    # ---------- Aba Liquidado_Mensal ----------
    dfl = pd.ExcelFile(SRC_LIQ).parse('comparativo 3', header=None)
    liq_rows = []
    for _, rr in dfl.iterrows():
        if isinstance(rr[0], str) and rr[0] in TARGETS:
            pt = rr[0]; ft = str(int(rr[1]))[:3]
            months = [float(rr[c]) if pd.notna(rr[c]) else 0.0 for c in range(2, 2 + NCLOSED)]
            liq_rows.append((pt, ft, months))
    def median4(vals):
        # mediana genérica; para 5 meses fechados (abr–ago) = valor central
        s = sorted(vals)
        if not s: return 0.0
        n = len(s)
        return s[n//2] if n % 2 else (s[n//2 - 1] + s[n//2]) / 2

    def median_ft(lr, ft):
        # mediana dos totais mensais (abr–ago) da fonte, somados entre linhas do mesmo mês
        tot = [sum(x[2][mi] for x in lr if x[1] == ft) for mi in range(WIN_START, NCLOSED)]
        return median4(tot)

    ml = wb.create_sheet('Liquidado_Mensal')
    ml.sheet_view.showGridLines = False
    ml['A1'] = f'LIQUIDADO MENSAL 2026 (base do Comparativo 2) — fonte: {LIQ_NAME}'
    ml['A1'].font = f(11, True, '1F4E78')
    mheads = ['PROGRAMA DE TRABALHO', 'FONTE (3 díg)'] + [m.upper() for m in _MN[:NCLOSED]] + \
             [f'ACUMULADO ({ACUM_LBL})', f'MÉDIA {WIN} MESES ({WIN_LBL})', f'MEDIANA {WIN} MESES ({WIN_LBL})']
    for c, h in enumerate(mheads, 1):
        cell = ml.cell(3, c, h)
        cell.font = f(9, True, 'FFFFFF'); cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ml.row_dimensions[3].height = 30
    mr = 4
    for pt, ft, months in liq_rows:
        ml.cell(mr, 1, pt).font = f(9); ml.cell(mr, 1).border = border
        fc = ml.cell(mr, 2, ft); fc.font = f(9); fc.border = border
        fc.alignment = Alignment(horizontal='center')
        for k, mv in enumerate(months):
            c = ml.cell(mr, 3 + k, mv)
            c.number_format = NUM; c.font = f(9); c.border = border
        ca = ml.cell(mr, 3 + NCLOSED, sum(months))
        ca.number_format = NUM; ca.font = f(9, True); ca.border = border
        cm = ml.cell(mr, 4 + NCLOSED, sum(months[WIN_START:NCLOSED]) / WIN)
        cm.number_format = NUM; cm.font = f(9, True); cm.border = border
        cmd = ml.cell(mr, 5 + NCLOSED, median4(months[WIN_START:NCLOSED]))
        cmd.number_format = NUM; cmd.font = f(9, True); cmd.border = border
        mr += 1
    mw = [22, 12] + [14]*NCLOSED + [20, 22, 22]
    for c, w in enumerate(mw, 1): ml.column_dimensions[get_column_letter(c)].width = w
    _win_meses = ','.join(_MN[WIN_START:NCLOSED])
    ml.cell(mr, 1, f'Obs.: "acumulado" = soma {ACUM_LBL}; "média {WIN} meses" = média de ({_win_meses}); "mediana {WIN} meses" = mediana de ({_win_meses}). Comparativo 2 usa a média e Comparativo 3 usa a mediana; projeção nas abas de PT = acumulado + (média ou mediana) × meses restantes ({REST_LBL} = {REST}).').font = f(8, False, '808080')
    ml.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=5 + NCLOSED)

    # ---------- Colunas de dados na QDD_Filtrado ----------
    Y25 = {'LEI': 'H', 'ALT': 'I', 'DAUT': 'L', 'EMP': 'M', 'LIQ': 'N'}
    Y26 = {'LEI': 'Q', 'ALT': 'R', 'COTA': 'S', 'CONT': 'T', 'DAUT': 'U', 'EMP': 'V', 'DISP': 'X'}
    FATOR = 'Consolidado!$B$5'
    MESES_REST = 'Consolidado!$B$6'

    def sumifs(col, pt, fonte=None):
        s = f"SUMIFS(QDD_Filtrado!{col}:{col},QDD_Filtrado!$C:$C,\"{pt}\""
        if fonte:
            s += f",QDD_Filtrado!$F:$F,\"{fonte}*\""
        return s + ")"

    def sumifs_liq(col, pt, fonte):
        return (f"SUMIFS(Liquidado_Mensal!{col}:{col},Liquidado_Mensal!$A:$A,\"{pt}\","
                f"Liquidado_Mensal!$B:$B,\"{fonte}\")")

    # ---------- Métricas por PT (para os pontos de atenção) ----------
    def g(row, col_letter):
        idx = openpyxl.utils.column_index_from_string(col_letter) - 1
        v = row[idx]
        return float(v) if isinstance(v, (int, float)) else 0.0

    # FAT e REST vêm do bloco de parâmetros (derivados de MESES_FECHADOS)
    metrics = {}
    for tab in TABS:
        if tab['group']:
            prows = [r for r in rows_data if action_of(r[2]) == tab['action']]
            lr = [x for x in liq_rows if action_of(x[0]) == tab['action']]
        else:
            prows = [r for r in rows_data if r[2] == tab['match']]
            lr = [x for x in liq_rows if x[0] == tab['match']]
        m = {}
        m['emp25'] = sum(g(r, 'M') for r in prows)
        m['liq25'] = sum(g(r, 'N') for r in prows)
        m['dot26'] = sum(g(r, 'Q') + g(r, 'R') for r in prows)
        m['daut26'] = sum(g(r, 'U') for r in prows)
        m['emp26'] = sum(g(r, 'V') for r in prows)
        m['cota26'] = sum(g(r, 'S') for r in prows)
        m['cont26'] = sum(g(r, 'T') for r in prows)
        m['disp26'] = sum(g(r, 'X') for r in prows)
        m['liq_acum'] = sum(sum(x[2]) for x in lr)
        m['liq_med'] = sum(sum(x[2][WIN_START:NCLOSED]) / WIN for x in lr)
        m['proj_emp'] = m['emp26'] * FAT
        m['proj_liq'] = m['liq_acum'] + m['liq_med'] * REST
        m['c1'] = m['dot26'] - m['emp25']
        m['c2'] = m['dot26'] - m['proj_emp']
        m['c3'] = m['dot26'] - m['proj_liq']
        m['med_liq'] = sum(median_ft(lr, ft) for ft in tab['fontes'])
        m['proj_liq_med'] = m['liq_acum'] + m['med_liq'] * REST
        m['c3_med'] = m['dot26'] - m['proj_liq_med']
        m['alt26'] = sum(g(r, 'R') for r in prows)
        m['pct_emp'] = (m['emp26'] / m['daut26']) if m['daut26'] else 0.0
        m['nfontes'] = len(tab['fontes'])
        metrics[tab['sheet']] = m

    # ---------- Cálculo em Python (valores estáticos, sem depender do recálculo do Excel) ----------
    C25 = {'LEI': 'H', 'ALT': 'I', 'DAUT': 'L', 'EMP': 'M', 'LIQ': 'N'}
    C26 = {'LEI': 'Q', 'ALT': 'R', 'COTA': 'S', 'CONT': 'T', 'DAUT': 'U', 'EMP': 'V', 'DISP': 'X'}

    def rows_for(tab):
        if tab['group']:
            return [r for r in rows_data if action_of(r[2]) == tab['action']]
        return [r for r in rows_data if r[2] == tab['match']]

    def liq_for(tab):
        if tab['group']:
            return [x for x in liq_rows if action_of(x[0]) == tab['action']]
        return [x for x in liq_rows if x[0] == tab['match']]

    def col_sum(prows, colletter, ft):
        return sum(g(r, colletter) for r in prows if str(r[5])[:3] == ft)

    def acum_ft(lr, ft):
        return sum(sum(x[2]) for x in lr if x[1] == ft)

    def med_ft(lr, ft):
        return sum(sum(x[2][WIN_START:NCLOSED]) / WIN for x in lr if x[1] == ft)

    def value_for(spec, prows, lr, ft):
        if isinstance(spec, tuple):
            y, k = spec
            return col_sum(prows, C25[k] if y == '25' else C26[k], ft)
        dot = col_sum(prows, 'Q', ft) + col_sum(prows, 'R', ft)
        if spec == 'CALC_DOT': return dot
        if spec == 'C1_DOT':   return dot - col_sum(prows, 'M', ft)
        if spec == 'L3_ACUM':  return acum_ft(lr, ft)
        if spec == 'L3_MED':   return med_ft(lr, ft)
        if spec == 'L3_PROJ':  return acum_ft(lr, ft) + med_ft(lr, ft) * REST
        if spec == 'C3_DOT':   return dot - (acum_ft(lr, ft) + med_ft(lr, ft) * REST)
        if spec == 'L3_MEDIAN':   return median_ft(lr, ft)
        if spec == 'L3_PROJ_MED': return acum_ft(lr, ft) + median_ft(lr, ft) * REST
        if spec == 'C3_MED_DOT':  return dot - (acum_ft(lr, ft) + median_ft(lr, ft) * REST)
        return 0.0

    # Fontes consideradas na aba consolidada filtrada
    ALLOWED_FONTES = {'100', '101', '102', '161', '178', '183'}

    def compute_metrics(allowed=None):
        md = {}
        for tab in TABS:
            prows = rows_for(tab); lr = liq_for(tab)
            if allowed is not None:
                prows = [r for r in prows if str(r[5])[:3] in allowed]
                lr = [x for x in lr if x[1] in allowed]
            m = {}
            m['emp25'] = sum(g(r, 'M') for r in prows)
            m['dot26'] = sum(g(r, 'Q') + g(r, 'R') for r in prows)
            m['emp26'] = sum(g(r, 'V') for r in prows)
            m['alt26'] = sum(g(r, 'R') for r in prows)
            acum = sum(sum(x[2]) for x in lr)
            med = sum(sum(x[2][WIN_START:NCLOSED]) / WIN for x in lr)
            m['proj_liq'] = acum + med * REST
            m['c1'] = m['dot26'] - m['emp25']
            m['c3'] = m['dot26'] - m['proj_liq']
            fts_here = sorted(set(x[1] for x in lr))
            m['med_liq'] = sum(median_ft(lr, ft) for ft in fts_here)
            m['proj_liq_med'] = acum + m['med_liq'] * REST
            m['c3_med'] = m['dot26'] - m['proj_liq_med']
            md[tab['sheet']] = m
        return md

    def rs(v):
        a = abs(v)
        if a >= 1e9: return f"R$ {v/1e9:,.2f} bi".replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {v/1e6:,.1f} mi".replace(',', 'X').replace('.', ',').replace('X', '.')

    def gerar_pontos(tab):
        m = metrics[tab['sheet']]
        b = []
        if tab.get('group'):
            b.append(f"Consolidação: esta aba agrega os programas da Ação {tab['action']} — "
                     + ", ".join(tab['members']) + ". Valores somados por fonte de recursos.")
        if tab['key'] == '08.244.6228.4272.0001':
            b.append("Transformação de PT: durante 2025 o programa 08.306.6228.4173.0003 foi transformado neste "
                     "(08.244.6228.4272.0001). A execução de 2025 abaixo consolida os dois códigos; em 2026 existe apenas o 4272.0001.")
        pct = f"{m['pct_emp']*100:.1f}".replace('.', ',')
        # 1. Comparação 1
        if m['c1'] < 0:
            b.append(f"Comparação 1: a dotação autorizada 2026 ({rs(m['dot26'])}) é inferior ao empenhado de 2025 "
                     f"({rs(m['emp25'])}), déficit potencial de {rs(abs(m['c1']))} caso a demanda repita o exercício anterior.")
        else:
            b.append(f"Comparação 1: a dotação autorizada 2026 ({rs(m['dot26'])}) supera o empenhado de 2025 "
                     f"({rs(m['emp25'])}), folga de {rs(m['c1'])} frente ao exercício anterior.")
        # 2. Ritmo de execução (empenho)
        b.append(f"Ritmo de execução: {pct}% da despesa autorizada 2026 já empenhada com {PCT_DEC}% do ano decorrido ({LAST_M} fechado).")
        # 3. Comparação 2
        if m['c3'] < 0:
            b.append(f"Comparação 2: a projeção do liquidado até dez ({rs(m['proj_liq'])}) supera a dotação — déficit de {rs(abs(m['c3']))}.")
        else:
            b.append(f"Comparação 2: a projeção do liquidado até dez ({rs(m['proj_liq'])}) fica abaixo da dotação — folga de {rs(m['c3'])}.")
        # 3b. Comparação 3 (mediana)
        if m['c3_med'] < 0:
            b.append(f"Comparação 3 (mediana): pela mediana dos {WIN} meses fechados ({WIN_LBL}), a projeção do liquidado até dez ({rs(m['proj_liq_med'])}) supera a dotação — déficit de {rs(abs(m['c3_med']))}.")
        else:
            b.append(f"Comparação 3 (mediana): pela mediana dos {WIN} meses fechados ({WIN_LBL}), a projeção do liquidado até dez ({rs(m['proj_liq_med'])}) fica abaixo da dotação — folga de {rs(m['c3_med'])}.")
        # 4. Cota / contingenciado
        if m['cota26'] > 1e6 or m['cont26'] > 1e6:
            partes = []
            if m['cota26'] > 1e6: partes.append(f"{rs(m['cota26'])} em cota a liberar")
            if m['cont26'] > 1e6: partes.append(f"{rs(m['cont26'])} contingenciado")
            b.append("Recursos retidos: " + " e ".join(partes) +
                     " — a folga/déficit acima depende da liberação desses valores.")
        return b

    ROWDEFS = [
        ('EXECUÇÃO 2025 (exercício encerrado)', 'SEC', None, False),
        ('Dotação inicial (LEI) 2025', ('25', 'LEI'), NUM, False),
        ('Alterações 2025', ('25', 'ALT'), NUM, False),
        ('Despesa autorizada 2025 (DESP AUT.)', ('25', 'DAUT'), NUM, False),
        ('Empenhado 2025', ('25', 'EMP'), NUM, True),
        ('Liquidado 2025', ('25', 'LIQ'), NUM, False),
        ('ORÇAMENTO 2026 (posição do arquivo)', 'SEC', None, False),
        ('Dotação inicial (LEI) 2026', ('26', 'LEI'), NUM, False),
        ('Alterações 2026', ('26', 'ALT'), NUM, False),
        ('Dotação autorizada 2026 (LEI + ALTERAÇÃO)', 'CALC_DOT', NUM, True),
        ('Cota a liberar 2026', ('26', 'COTA'), NUM, False),
        ('Contingenciado 2026', ('26', 'CONT'), NUM, False),
        ('Despesa autorizada 2026 (DESP AUT.)', ('26', 'DAUT'), NUM, True),
        (f'Empenhado 2026 ({ACUM_LBL} fechado)', ('26', 'EMP'), NUM, True),
        ('Disponível 2026 (saldo)', ('26', 'DISP'), NUM, False),
        ('COMPARAÇÃO 1 — EMPENHO 2025 × ORÇAMENTO 2026', 'SEC', None, False),
        ('Déficit/superávit vs dotação autorizada (LEI+ALT)', 'C1_DOT', NUM, True),
        (f'COMPARAÇÃO 2 — PROJEÇÃO DO LIQUIDADO (média {WIN_LBL}, {WIN} meses)', 'SEC', None, False),
        (f'Liquidado acumulado 2026 ({ACUM_LBL})', 'L3_ACUM', NUM, False),
        (f'Média mensal do liquidado ({WIN_LBL})', 'L3_MED', NUM, False),
        ('Projeção do liquidado até dez (acum + média × meses restantes)', 'L3_PROJ', NUM, False),
        ('Déficit/superávit vs dotação autorizada (LEI+ALT)', 'C3_DOT', NUM, True),
        (f'COMPARAÇÃO 3 — PROJEÇÃO DO LIQUIDADO (mediana {WIN_LBL}, {WIN} meses)', 'SEC', None, False),
        (f'Liquidado acumulado 2026 ({ACUM_LBL})', 'L3_ACUM', NUM, False),
        (f'Mediana mensal do liquidado ({WIN_LBL})', 'L3_MEDIAN', NUM, False),
        ('Projeção do liquidado até dez (acum + mediana × meses restantes)', 'L3_PROJ_MED', NUM, False),
        ('Déficit/superávit vs dotação autorizada (LEI+ALT)', 'C3_MED_DOT', NUM, True),
    ]
    START = 5
    def rowno(n): return START + n
    R_EMP25 = rowno(4)
    R_DOT26 = rowno(9)
    R_DAUT26 = rowno(12)
    R_EMP26 = rowno(13)
    R_L3_ACUM = rowno(18)
    R_L3_MED = rowno(19)
    R_L3_PROJ = rowno(20)

    for tab in TABS:
        pt = tab['match']
        fts = tab['fontes']
        prows = rows_for(tab); lr = liq_for(tab)
        sh = wb.create_sheet(tab['sheet'])
        sh.sheet_view.showGridLines = False
        total_col = 2 + len(fts)
        sh['A1'] = tab['desc'] if tab['group'] else f"{tab['key']} — {tab['desc']}"
        sh['A1'].font = f(12, True, '1F4E78')
        subt = 'Consolida os PTs de Passe Livre da Ação 4202 | ' if tab['group'] else ''
        sh['A2'] = subt + 'Segregação por fonte de recursos (3 primeiros dígitos) | Valores em R$ | Negativo (vermelho) = déficit'
        sh['A2'].font = f(9, False, '808080')
        hc = sh.cell(4, 1, 'INDICADOR')
        hc.font = f(9, True, 'FFFFFF'); hc.fill = hfill; hc.border = border
        for j, ft in enumerate(fts):
            c = sh.cell(4, 2 + j, f'FONTE {ft}')
            c.font = f(9, True, 'FFFFFF'); c.fill = hfill; c.border = border
            c.alignment = Alignment(horizontal='center')
        c = sh.cell(4, total_col, 'TOTAL')
        c.font = f(9, True, 'FFFFFF'); c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal='center')

        r = START
        for lab, spec, fmt, bold in ROWDEFS:
            lc = sh.cell(r, 1, lab)
            if spec == 'SEC':
                lc.font = f(9, True, '1F4E78'); lc.fill = gfill
                for cl in range(2, total_col + 1): sh.cell(r, cl).fill = gfill
            else:
                lc.font = f(9, bold); lc.border = border
                for j, ft in enumerate(fts):
                    cell = sh.cell(r, 2 + j, value_for(spec, prows, lr, ft))
                    cell.number_format = fmt; cell.font = f(9, bold); cell.border = border
                tot_val = sum(value_for(spec, prows, lr, ft) for ft in fts)
                tc = sh.cell(r, total_col, tot_val)
                tc.number_format = fmt; tc.font = f(9, True); tc.border = border
                tc.fill = h2fill
            r += 1

        # ---------- Pontos de Atenção ----------
        r += 1
        hc = sh.cell(r, 1, 'PONTOS DE ATENÇÃO')
        hc.font = f(10, True, 'FFFFFF'); hc.fill = hfill
        hc.alignment = Alignment(horizontal='left', vertical='center')
        for cl in range(2, total_col + 1):
            sh.cell(r, cl).fill = hfill
        sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_col)
        r += 1
        for i, ponto in enumerate(gerar_pontos(tab), 1):
            cell = sh.cell(r, 1, f'{i}.  {ponto}')
            cell.font = f(9); cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = gfill
            sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_col)
            nlin = max(1, (len(ponto) // 110) + 1)
            sh.row_dimensions[r].height = 15 * nlin + 3
            r += 1

        # ---------- Dados e gráficos ----------
        tcol = get_column_letter(total_col)
        anchor_col = get_column_letter(total_col + 2)
        mtab = metrics[tab['sheet']]
        r += 1
        cd = r
        sh.cell(cd, 1, 'DADOS — GRÁFICO 1 (Comparação 1)').font = f(9, True, '1F4E78')
        sh.cell(cd + 1, 1, 'Empenhado 2025').font = f(9)
        c = sh.cell(cd + 1, 2, mtab['emp25']); c.number_format = NUM; c.font = f(9)
        sh.cell(cd + 2, 1, 'Dotação Autorizada 2026').font = f(9)
        c = sh.cell(cd + 2, 2, mtab['dot26']); c.number_format = NUM; c.font = f(9)

        bar = BarChart(); bar.type = 'col'; bar.varyColors = True; bar.legend = None
        bar.title = 'Comparação 1 — Empenho 2025 × Dotação 2026 (todas as fontes)'
        bar.y_axis.title = 'R$'; bar.height = 7.2; bar.width = 13
        bdata = Reference(sh, min_col=2, min_row=cd + 1, max_row=cd + 2)
        bcats = Reference(sh, min_col=1, min_row=cd + 1, max_row=cd + 2)
        bar.add_data(bdata, titles_from_data=False); bar.set_categories(bcats)
        sh.add_chart(bar, f'{anchor_col}{cd}')

        # Dados do gráfico 2: liquidado acumulado mês a mês + dotação (linha fixa)
        lh = cd + 4
        sh.cell(lh, 1, 'Mês').font = f(9, True)
        sh.cell(lh, 2, 'Liquidado acumulado').font = f(9, True)
        sh.cell(lh, 3, 'Dotação autorizada 2026').font = f(9, True)
        ML = [m.capitalize() for m in _MN[:NCLOSED]]
        mtot = [sum(x[2][mi] for x in lr) for mi in range(NCLOSED)]
        cum = 0.0
        for i, mlab in enumerate(ML):
            rr = lh + 1 + i
            cum += mtot[i]
            sh.cell(rr, 1, mlab).font = f(9)
            cell = sh.cell(rr, 2, cum); cell.number_format = NUM; cell.font = f(9)
            c = sh.cell(rr, 3, mtab['dot26']); c.number_format = NUM; c.font = f(9)

        line = LineChart(); line.title = 'Comparação 2 — Liquidado acum. × Dotação (todas as fontes)'
        line.y_axis.title = 'R$'; line.height = 7.2; line.width = 13
        ldata = Reference(sh, min_col=2, max_col=3, min_row=lh, max_row=lh + NCLOSED)
        line.add_data(ldata, titles_from_data=True)
        lcats = Reference(sh, min_col=1, min_row=lh + 1, max_row=lh + NCLOSED)
        line.set_categories(lcats)
        sh.add_chart(line, f'{anchor_col}{cd + 16}')

        # ---------- Gráficos FONTES TESOURO (100,101,102,161,178,183) ----------
        prows_t = [rr_ for rr_ in prows if str(rr_[5])[:3] in ALLOWED_FONTES]
        lr_t = [x for x in lr if x[1] in ALLOWED_FONTES]
        emp25_t = sum(g(rr_, 'M') for rr_ in prows_t)
        dot26_t = sum(g(rr_, 'Q') + g(rr_, 'R') for rr_ in prows_t)

        cdt = lh + NCLOSED + 1
        sh.cell(cdt, 1, 'DADOS — GRÁFICO 1 (Comp. 1 — FONTES TESOURO)').font = f(9, True, 'C55A11')
        sh.cell(cdt + 1, 1, 'Empenhado 2025 (tesouro)').font = f(9)
        c = sh.cell(cdt + 1, 2, emp25_t); c.number_format = NUM; c.font = f(9)
        sh.cell(cdt + 2, 1, 'Dotação Autorizada 2026 (tesouro)').font = f(9)
        c = sh.cell(cdt + 2, 2, dot26_t); c.number_format = NUM; c.font = f(9)

        bar_t = BarChart(); bar_t.type = 'col'; bar_t.varyColors = True; bar_t.legend = None
        bar_t.title = 'Comparação 1 — Empenho 2025 × Dotação 2026 (fontes tesouro)'
        bar_t.y_axis.title = 'R$'; bar_t.height = 7.2; bar_t.width = 13
        bdata = Reference(sh, min_col=2, min_row=cdt + 1, max_row=cdt + 2)
        bcats = Reference(sh, min_col=1, min_row=cdt + 1, max_row=cdt + 2)
        bar_t.add_data(bdata, titles_from_data=False); bar_t.set_categories(bcats)
        sh.add_chart(bar_t, f'{anchor_col}{cd + 32}')

        lht = cdt + 4
        sh.cell(lht, 1, 'Mês').font = f(9, True)
        sh.cell(lht, 2, 'Liquidado acum. (tesouro)').font = f(9, True)
        sh.cell(lht, 3, 'Dotação 2026 (tesouro)').font = f(9, True)
        mtot_t = [sum(x[2][mi] for x in lr_t) for mi in range(NCLOSED)]
        cumt = 0.0
        for i, mlab in enumerate(ML):
            rr = lht + 1 + i
            cumt += mtot_t[i]
            sh.cell(rr, 1, mlab).font = f(9)
            c = sh.cell(rr, 2, cumt); c.number_format = NUM; c.font = f(9)
            c = sh.cell(rr, 3, dot26_t); c.number_format = NUM; c.font = f(9)

        line_t = LineChart(); line_t.title = 'Comparação 2 — Liquidado acum. × Dotação (fontes tesouro)'
        line_t.y_axis.title = 'R$'; line_t.height = 7.2; line_t.width = 13
        ldata = Reference(sh, min_col=2, max_col=3, min_row=lht, max_row=lht + NCLOSED)
        line_t.add_data(ldata, titles_from_data=True)
        lcats = Reference(sh, min_col=1, min_row=lht + 1, max_row=lht + NCLOSED)
        line_t.set_categories(lcats)
        sh.add_chart(line_t, f'{anchor_col}{cd + 48}')

        sh.column_dimensions['A'].width = 50
        for j in range(2, total_col + 1):
            sh.column_dimensions[get_column_letter(j)].width = 16
        sh.freeze_panes = 'B5'

    # ---------- Consolidado (função reutilizável) ----------
    def build_consolidado(name, md, allowed, subtitle, extra_notes):
      co = wb.create_sheet(name)
      co.sheet_view.showGridLines = False
      co['A1'] = 'DÉFICIT CONSOLIDADO — POSSÍVEIS DÉFICITS ATÉ O ENCERRAMENTO DE 2026'
      co['A1'].font = f(13, True, '1F4E78')
      co['A2'] = subtitle
      co['A2'].font = f(9, False, '808080')
      co['A4'] = f'Meses fechados de execução 2026 ({ACUM_LBL})'; co['A4'].font = f(9)
      co['B4'] = NCLOSED; co['B4'].number_format = '0'
      co['A5'] = 'Fator de anualização (12/meses fechados)'; co['A5'].font = f(9)
      co['B5'] = FAT; co['B5'].number_format = '0.000'
      co['A6'] = f'Meses restantes {REST_LBL} (Comp. 2 e 3)'; co['A6'].font = f(9)
      co['B6'] = REST; co['B6'].number_format = '0'
      for rr in (4, 5, 6):
          c = co.cell(rr, 2)
          c.font = Font(name=F, size=9, color='0000FF'); c.fill = yfill; c.border = border
      co['C4'] = f'Premissa informada pelo usuário ({LAST_M} fechado)'; co['C4'].font = f(8, False, '808080')

      heads = ['PROGRAMA DE TRABALHO', 'DESCRIÇÃO',
               'EMPENHADO 2025', 'DOTAÇÃO AUTORIZADA 2026 (LEI+ALT)',
               f'EMPENHADO 2026 ({ACUM_LBL.upper()})',
               f'PROJEÇÃO LIQUIDADO 2026 (MÉDIA {WIN}M {WIN_LBL})',
               f'PROJEÇÃO LIQUIDADO 2026 (MEDIANA {WIN}M {WIN_LBL})',
               'COMP. 1 — DÉFICIT: EMPENHO 2025 × DOTAÇÃO 2026',
               'ALTERAÇÃO 2026 (créditos add./cancel.)',
               'COMP. 2 — DÉFICIT: PROJEÇÃO DO LIQUIDADO 2026 × DOTAÇÃO',
               'COMP. 3 — DÉFICIT: PROJEÇÃO (MEDIANA) × DOTAÇÃO']
      HR = 8
      for j, h in enumerate(heads, 1):
          c = co.cell(HR, j, h)
          c.font = f(8, True, 'FFFFFF'); c.fill = hfill; c.border = border
          c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      co.row_dimensions[HR].height = 40

      r = HR + 1
      for tab in TABS:
          m = md[tab['sheet']]
          co.cell(r, 1, tab['key']).font = f(9)
          co.cell(r, 2, tab['desc']).font = f(8)
          vals = [m['emp25'], m['dot26'], m['emp26'], m['proj_liq'], m['proj_liq_med'], m['c1'], m['alt26'], m['c3'], m['c3_med']]
          for j, v in enumerate(vals, 3):
              c = co.cell(r, j, v)
              c.number_format = NUM; c.font = f(9); c.border = border
          co.cell(r, 1).border = border; co.cell(r, 2).border = border
          r += 1
      co.cell(r, 1, f'TOTAL ({len(TABS)} linhas — {len(TARGETS)} PTs)').font = f(9, True)
      co.cell(r, 1).border = border; co.cell(r, 2).border = border
      co.cell(r, 1).fill = h2fill; co.cell(r, 2).fill = h2fill
      NCOL = len(heads)
      _keys = ['emp25', 'dot26', 'emp26', 'proj_liq', 'proj_liq_med', 'c1', 'alt26', 'c3', 'c3_med']
      for j in range(3, NCOL + 1):
          tot = sum(md[t['sheet']][_keys[j - 3]] for t in TABS)
          c = co.cell(r, j, tot)
          c.number_format = NUM; c.font = f(9, True); c.border = border; c.fill = h2fill

      # Destaque das colunas H (Comp.1) e I (Alteração 2026) — moldura única
      hl = Side(style='medium', color='C55A11')
      non = Side(style=None)
      for rr in range(HR, r + 1):
          top = hl if rr == HR else border.top
          bot = hl if rr == r else border.bottom
          co.cell(rr, 8).border = Border(left=hl, right=non, top=top, bottom=bot)
          co.cell(rr, 9).border = Border(left=non, right=hl, top=top, bottom=bot)

      notes = [
       'NOTAS',
       f'1. Fonte: "{QDD_NAME}" (2025 e 2026) e "{LIQ_NAME}" (liquidado mensal). Marcadores originais na aba QDD_Filtrado; abas por PT segregam por fonte de recursos (3 primeiros dígitos).',
       f'2. Comparação 2 (liquidado — média): projeção = liquidado acumulado ({ACUM_LBL}) + média mensal dos {WIN} meses fechados ({WIN_LBL}) × meses restantes ({REST_LBL} = {REST}). Comparação 3 (liquidado — mediana): mesma fórmula, trocando a média pela mediana dos mesmos {WIN} meses ({WIN_LBL}) — mais robusta a meses atípicos. A mediana é calculada por fonte sobre os totais mensais e somada entre fontes. Último mês tratado como fechado: {LAST_M} ({NCLOSED} meses). Base na aba Liquidado_Mensal.',
       '3. Consolidações por ação: (a) Passe Livre (Ação 4202) — 26.453.6216.4202.0002 (PNE), 12.361.6221.4202.0004 e 12.453.6221.4202.0010 (Estudantil) agregados em PT_4202_Consolidado; (b) Equilíbrio Financeiro do STPC (Ação 2455) — 26.453.6216.2455.0002 e 26.453.6216.2455.0001 agregados em PT_2455_Consolidado. Valores somados por fonte de recursos.',
       '4. Transformação de PT: o programa 08.306.6228.4173.0003 foi transformado em 08.244.6228.4272.0001 durante 2025. A execução de 2025 dos dois códigos foi consolidada sob o 4272.0001 (Cartão Prato Cheio); em 2026 existe apenas o 4272.0001.',
       '5. Valores fixos (snapshot): os números das comparações foram gravados como valores calculados a partir dos arquivos-fonte, e não como fórmulas — assim a planilha não recalcula (nem altera os números) ao ser aberta/editada no Excel. Para atualizar, é preciso reprocessar com novos arquivos.',
       '6. Convenção: valores negativos (em vermelho) = déficit; positivos = superávit/folga.',
      ] + list(extra_notes)
      nr = r + 2
      for i, n in enumerate(notes):
          c = co.cell(nr + i, 1, n)
          c.font = Font(name=F, size=8, bold=(i == 0), color='1F4E78' if i == 0 else '404040')
          c.alignment = Alignment(wrap_text=True, vertical='top')
          co.merge_cells(start_row=nr+i, start_column=1, end_row=nr+i, end_column=NCOL)
          co.row_dimensions[nr+i].height = 12 if i == 0 else 24

      co.column_dimensions['A'].width = 22
      co.column_dimensions['B'].width = 34
      for j in range(3, NCOL + 1): co.column_dimensions[get_column_letter(j)].width = 18

      # Gráficos
      PT_FIRST = HR + 1
      PT_LAST = r - 1
      note_end = nr + len(notes) - 1
      bar = BarChart(); bar.type = 'col'; bar.grouping = 'clustered'
      bar.title = 'Empenhado 2025 × Dotação Autorizada 2026 — por programa'
      bar.y_axis.title = 'R$'; bar.height = 9.5; bar.width = 26
      bdata = Reference(co, min_col=3, max_col=4, min_row=HR, max_row=PT_LAST)
      bar.add_data(bdata, titles_from_data=True)
      bcats = Reference(co, min_col=2, min_row=PT_FIRST, max_row=PT_LAST)
      bar.set_categories(bcats)
      co.add_chart(bar, f'A{note_end + 2}')

      ch = note_end + 22
      co.cell(ch, 1, 'DADOS — GRÁFICO DE LINHA (Consolidado)').font = f(9, True, '1F4E78')
      lh = ch + 1
      co.cell(lh, 1, 'Mês').font = f(9, True)
      co.cell(lh, 2, 'Liquidado acumulado (total)').font = f(9, True)
      co.cell(lh, 3, 'Dotação autorizada total 2026').font = f(9, True)
      ML = [m.capitalize() for m in _MN[:NCLOSED]]
      allmtot = [sum(x[2][mi] for x in liq_rows if (allowed is None or x[1] in allowed)) for mi in range(NCLOSED)]
      dot_total = sum(md[t['sheet']]['dot26'] for t in TABS)
      cum = 0.0
      for i, mlab in enumerate(ML):
          rr = lh + 1 + i
          cum += allmtot[i]
          co.cell(rr, 1, mlab).font = f(9)
          cell = co.cell(rr, 2, cum); cell.number_format = NUM; cell.font = f(9)
          c = co.cell(rr, 3, dot_total); c.number_format = NUM; c.font = f(9)

      line = LineChart(); line.title = 'Liquidado acumulado (total) × Dotação total 2026'
      line.y_axis.title = 'R$'; line.height = 9; line.width = 20
      ldata = Reference(co, min_col=2, max_col=3, min_row=lh, max_row=lh + NCLOSED)
      line.add_data(ldata, titles_from_data=True)
      lcats = Reference(co, min_col=1, min_row=lh + 1, max_row=lh + NCLOSED)
      line.set_categories(lcats)
      co.add_chart(line, f'E{ch}')

    # Aba consolidada principal (todas as fontes)
    build_consolidado('Consolidado', metrics, None,
        f'Base: {QDD_NAME} + {LIQ_NAME} | Valores em R$ | Negativo (vermelho) = déficit',
        [])
    # Aba consolidada filtrada (somente fontes selecionadas)
    build_consolidado('Consolidado (Fontes 100-183)', compute_metrics(ALLOWED_FONTES), ALLOWED_FONTES,
        f'SOMENTE fontes 100, 101, 102, 161, 178 e 183 | Base: {QDD_NAME} | Valores em R$ | Negativo (vermelho) = déficit',
        ['7. Filtro de fontes: esta aba considera, em TODOS os cálculos (empenho, dotação, liquidado e projeções), APENAS as fontes 100, 101, 102, 161, 178 e 183. Todas as demais fontes foram excluídas.'])

    wb.save(OUT)
    print('saved; abas:', wb.sheetnames)

# =============================== PARSER DO DASHBOARD =========================================
TREAS = ['100', '101', '102', '161', '178', '183']  # fontes do tesouro
HERE = os.path.dirname(os.path.abspath(__file__))
MONTHS = {'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'}


def rows(ws):
    out = []
    for r in ws.iter_rows(values_only=True):
        row = list(r)
        while row and row[-1] is None:
            row.pop()
        out.append(row)
    return out


def num(v):
    return v if isinstance(v, (int, float)) else 0


def parse_consol(wb, name):
    cr = rows(wb[name])
    hdr_i = next(i for i, r in enumerate(cr) if r and r[0] == 'PROGRAMA DE TRABALHO')
    hdr = [str(c) if c is not None else '' for c in cr[hdr_i]]

    def col(*preds, default=None):
        # acha o índice da 1ª coluna cujo cabeçalho satisfaz todos os predicados (substrings, case-insensitive)
        for idx, h in enumerate(hdr):
            hu = h.upper()
            if all(p.upper() in hu for p in preds):
                return idx
        return default
    # Layout lido por NOME de cabeçalho (robusto à ordem e a colunas novas: mediana/Comp.3)
    IX = {
        'emp25': col('EMPENHADO 2025', default=2),
        'dot26': col('DOTAÇÃO AUTORIZADA', default=3),
        'emp26': col('EMPENHADO 2026', default=4),
        'proj_liq': col('PROJEÇÃO LIQUIDADO', 'MÉDIA', default=col('PROJEÇÃO LIQUIDADO', default=5)),
        'proj_liq_med': col('PROJEÇÃO LIQUIDADO', 'MEDIANA'),
        'comp1': col('COMP. 1', default=6),
        'alt26': col('ALTERAÇÃO 2026', default=7),
        'comp2': col('COMP. 2', default=8),
        'comp3': col('COMP. 3'),
    }
    progs, total, total_label = [], None, ''
    for r in cr[hdr_i + 1:]:
        if not r or r[0] is None:
            continue
        if str(r[0]).startswith('TOTAL'):
            total = r
            total_label = str(r[0])
            continue
        if str(r[0]).startswith('NOTAS'):
            break
        progs.append(r)
    li = next(i for i, r in enumerate(cr) if r and r[0] == 'Mês')
    line = []
    for r in cr[li + 1:]:
        if not r or r[0] not in MONTHS or len(r) < 3:
            break
        line.append({'mes': r[0], 'liq': r[1], 'dot': r[2]})
    meses = next((r[1] for r in cr if r and isinstance(r[0], str)
                  and r[0].startswith('Meses fechados')), 7)
    base_note = next((str(r[0]).split('|')[0].replace('Base:', '').strip()
                      for r in cr if r and isinstance(r[0], str) and r[0].startswith('Base:')), '')

    def gv(r, idx):
        return r[idx] if (idx is not None and idx < len(r)) else None

    def mk(r):
        return {'code': r[0], 'desc': r[1],
                'emp25': gv(r, IX['emp25']), 'dot26': gv(r, IX['dot26']), 'emp26': gv(r, IX['emp26']),
                'proj_liq': gv(r, IX['proj_liq']), 'proj_liq_med': gv(r, IX['proj_liq_med']),
                'comp1': gv(r, IX['comp1']), 'alt26': gv(r, IX['alt26']),
                'comp2': gv(r, IX['comp2']), 'comp3': gv(r, IX['comp3'])}

    return {'meses': meses, 'total': mk(total), 'line': line,
            'progs': [mk(r) for r in progs], 'total_label': total_label, 'base_note': base_note}


def sheet_for(wb, code):
    m = re.search(r'\.(\d{4}\.\d{4})$', str(code))
    if m:
        k = 'PT_' + m.group(1).replace('.', '_')
        if k in wb.sheetnames:
            return k
    return None


OVERRIDE = {'Ação 4202 (Passe Livre)': 'PT_4202_Consolidado',
            'Ação 2455 (Equilíbrio STPC)': 'PT_2455_Consolidado'}


def parse_pt(wb, name):
    r = rows(wb[name])
    hi = next(i for i, x in enumerate(r) if x and x[0] == 'INDICADOR')
    fontes = [c for c in r[hi][1:] if c is not None]

    def vals(label):
        for x in r:
            if x and isinstance(x[0], str) and x[0].strip().startswith(label):
                return x[1:]
        return None

    labels = {'emp_25': 'Empenhado 2025', 'dot_aut_26': 'Dotação autorizada 2026',
              'emp_26': 'Empenhado 2026', 'liq_acum': 'Liquidado acumulado 2026',
              'proj': 'Projeção do liquidado até dez'}
    ind = {k: vals(l) for k, l in labels.items()}
    cota = vals('Cota a liberar 2026')
    conting = vals('Contingenciado 2026')
    defrows = [x for x in r if x and isinstance(x[0], str)
               and x[0].strip().startswith('Déficit/superávit vs dotação')]
    comp1_f = defrows[0][1:] if len(defrows) > 0 else None
    comp2_f = defrows[1][1:] if len(defrows) > 1 else None
    pi = next((i for i, x in enumerate(r) if x and x[0] == 'PONTOS DE ATENÇÃO'), None)
    pontos = []
    if pi is not None:
        for x in r[pi + 1:]:
            if not x or x[0] is None:
                break
            if isinstance(x[0], str) and x[0].startswith('DADOS'):
                break
            pontos.append(x[0])
    mis = [i for i, x in enumerate(r) if x and x[0] == 'Mês']

    def mb(idx):
        out = []
        for x in r[idx + 1:]:
            if not x or x[0] not in MONTHS or len(x) < 3:
                break
            out.append({'mes': x[0], 'liq': x[1], 'dot': x[2]})
        return out

    return {'fontes': fontes, 'ind': ind, 'cota': cota, 'conting': conting,
            'comp1_f': comp1_f, 'comp2_f': comp2_f, 'pontos': pontos,
            'mon_tot': mb(mis[0]) if len(mis) > 0 else [],
            'mon_tes': mb(mis[1]) if len(mis) > 1 else []}


def build_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    CT = parse_consol(wb, 'Consolidado')
    CS = parse_consol(wb, 'Consolidado (Fontes 100-183)')
    programs = []
    for i, pt in enumerate(CT['progs']):
        ps = CS['progs'][i]
        assert pt['code'] == ps['code'], (pt['code'], ps['code'])
        sheet = OVERRIDE.get(pt['code']) or sheet_for(wb, pt['code'])
        if not sheet:
            raise RuntimeError('Aba do PT não encontrada para ' + str(pt['code']))
        d = parse_pt(wb, sheet)
        cota = d['cota'] or []
        conting = d['conting'] or []
        fontes = d['fontes']
        cota_t = num(cota[-1] if cota else 0)
        cont_t = num(conting[-1] if conting else 0)
        cota_s = cont_s = 0
        for idx, fn in enumerate(fontes):
            m = re.search(r'(\d{3})', str(fn))
            if m and m.group(1) in TREAS:
                cota_s += num(cota[idx] if idx < len(cota) else 0)
                cont_s += num(conting[idx] if idx < len(conting) else 0)

        def metrics(src, monthly, cv, cc):
            return {'emp25': src['emp25'], 'dot26': src['dot26'], 'emp26': src['emp26'],
                    'proj_liq': src['proj_liq'], 'proj_liq_med': src.get('proj_liq_med'),
                    'comp1': src['comp1'], 'alt26': src['alt26'],
                    'comp2': src['comp2'], 'comp3': src.get('comp3'),
                    'monthly': monthly, 'cota': cv, 'conting': cc, 'retido': cv + cc}

        programs.append({'code': pt['code'], 'desc': pt['desc'],
                         'total': metrics(pt, d['mon_tot'], cota_t, cont_t),
                         'tesouro': metrics(ps, d['mon_tes'], cota_s, cont_s),
                         'pontos': d['pontos'], 'fontes': d['fontes'], 'ind': d['ind'],
                         'cota_f': d['cota'], 'conting_f': d['conting'],
                         'comp1_f': d['comp1_f'], 'comp2_f': d['comp2_f']})
    for key, src in (('total', CT), ('tesouro', CS)):
        src['total']['retido'] = sum(p[key]['retido'] for p in programs)
        src['total']['cota'] = sum(p[key]['cota'] for p in programs)
        src['total']['conting'] = sum(p[key]['conting'] for p in programs)

    # reconciliação (sanidade)
    for v in ('total', 'tesouro'):
        s1 = sum(p[v]['comp1'] for p in programs)
        s2 = sum(p[v]['comp2'] for p in programs)
        T = (CT if v == 'total' else CS)['total']
        assert abs(s1 - T['comp1']) < 1 and abs(s2 - T['comp2']) < 1, \
            'Somatório não reconcilia na visão ' + v

    data = {'meses_fechados': CT['meses'], 'treasury': TREAS,
            'views': {'total': {'total': CT['total'], 'line': CT['line']},
                      'tesouro': {'total': CS['total'], 'line': CS['line']}},
            'programs': programs}

    # meta p/ placeholders
    m = re.search(r'(\d+)\s*linhas\s*[—-]\s*(\d+)\s*PT', CT['total_label'])
    linhas = m.group(1) if m else str(len(programs))
    pts = m.group(2) if m else str(len(programs))
    meta = {'base_note': CT['base_note'] or 'planilha de análise',
            'linhas': linhas, 'pts': pts}
    return data, meta


# =============================== PAGINA (index.html) =========================================
def gerar_pagina(xlsx_path, out_html="index.html"):
    """Gera a pagina visual (dashboard) a partir da planilha do dia. Modo ONLINE (Chart.js via CDN)."""
    here = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(here, "template.html")
    data, meta = build_data(xlsx_path)
    tpl = open(template_path, encoding="utf-8").read()
    html = tpl.replace("/*__DATA__*/", json.dumps(data, ensure_ascii=False))
    _MN = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    mf = int(data['meses_fechados']) if str(data['meses_fechados']).isdigit() else 7
    last = _MN[mf-1]; nxt = _MN[mf] if mf < 12 else 'dez'
    html = (html.replace("{{BASE_NOTE}}", meta['base_note'])
                .replace("{{LINHAS}}", meta['linhas'])
                .replace("{{PTS}}", meta['pts'])
                .replace("{{ACUM_UP}}", f"Jan\u2013{last.capitalize()}")
                .replace("{{ACUM_LBL}}", f"jan\u2013{last}")
                .replace("{{PROJ_LBL}}", f"{nxt}\u2013dez")
                .replace("{{MESES_N}}", str(mf))
                .replace("{{WIN}}", str(mf-3))
                .replace("{{XLSX}}", os.path.basename(xlsx_path))
                .replace("{{ATUALIZADO}}", date.today().strftime("%d/%m/%Y")))
    with open(out_html, "w", encoding="utf-8") as f:
        f.write(html)
    return out_html


# =============================== CONEXAO / CONSULTAS / ORQUESTRACAO ===========================
def ler_sql(nome):
    for enc in ('utf-8', 'latin-1'):
        try:
            return open(nome, encoding=enc).read().strip().rstrip(';').strip()
        except UnicodeDecodeError:
            continue
    raise RuntimeError("Nao consegui ler " + nome)


def consultar(con, sql, colnames):
    cur = con.cursor()
    cur.execute(sql)
    dados = cur.fetchall()
    cur.close()
    return pd.DataFrame(dados, columns=colnames)


def iniciar_oracle_client():
    if INSTANT_CLIENT and os.path.isdir(INSTANT_CLIENT):
        try:
            oracledb.init_oracle_client(lib_dir=INSTANT_CLIENT); print("     Instant Client:", INSTANT_CLIENT); return
        except Exception:
            pass
    bases = [r"C:\oracle", r"C:\Oracle", r"C:\app", os.path.expanduser("~"),
             os.getcwd(), os.path.dirname(os.path.abspath(__file__)), "C:\\"]
    for base in bases:
        try:
            nomes = sorted(os.listdir(base), reverse=True)
        except Exception:
            continue
        for nome in nomes:
            if nome.lower().startswith("instantclient"):
                p = os.path.join(base, nome)
                try:
                    oracledb.init_oracle_client(lib_dir=p); print("     Instant Client encontrado em:", p); return
                except Exception:
                    pass
    try:
        oracledb.init_oracle_client(); print("     Instant Client via PATH"); return
    except Exception as e:
        sys.exit("Nao encontrei o Instant Client. Ajuste a linha INSTANT_CLIENT no topo. Detalhe: " + str(e))


def main():
    senha = os.environ.get("ORACLE_PWD")
    if not senha:
        sys.exit("Defina a senha antes de rodar:  set ORACLE_PWD=suasenha")
    if os.path.exists(SAIDA):
        try:
            with open(SAIDA, "a"):
                pass
        except PermissionError:
            sys.exit(f"O arquivo '{SAIDA}' esta ABERTO (no Excel?). Feche-o e rode de novo.")

    print("1/4  Conectando ao Oracle...")
    iniciar_oracle_client()
    dsn = oracledb.makedsn(HOST, PORT, sid=SID)
    with oracledb.connect(user=USER, password=senha, dsn=dsn) as con:
        print("2/4  Rodando as consultas (QDD e LIQUIDADO)...")
        dfq = consultar(con, ler_sql(SQL_QDD_FILE), QDD_COLS)
        dfl = consultar(con, ler_sql(SQL_LIQ_FILE), LIQ_COLS)
    print(f"     QDD: {len(dfq)} linhas | LIQUIDADO: {len(dfl)} linhas")

    for c in ['COEXERCICIO', 'COFONTEMAE', 'CONATUREZA', 'IDUSO', 'DOTMAISALT', 'DESPAUT',
              'ALT', 'LEI', 'CONT105', 'LIQ', 'EMP', 'DISP', 'CONT104']:
        dfq[c] = pd.to_numeric(dfq[c], errors='coerce').fillna(0.0)
    dfq['PT'] = dfq['PT'].astype(str)
    for c in ['COFONTEMAE', 'INMES', 'LIQ']:
        dfl[c] = pd.to_numeric(dfl[c], errors='coerce').fillna(0.0)
    dfl['PT'] = dfl['PT'].astype(str)

    print("3/4  Montando os arquivos-fonte...")
    os.makedirs("_fontes", exist_ok=True)
    qpath = os.path.join("_fontes", "qdd.xlsx")
    lpath = os.path.join("_fontes", "liq.xlsx")
    salvar_aba(montar_qdd(dfq), qpath, "2025 e 2026")
    salvar_aba(montar_liquidado(dfl), lpath, "comparativo 3")

    print("4/4  Gerando a projecao...")
    gerar_analise(qpath, lpath, SAIDA)

    print("5/5  Gerando a pagina (index.html)...")
    gerar_pagina(SAIDA, "index.html")
    print(f"\nPRONTO! Planilha: {SAIDA}  |  Pagina: index.html")


if __name__ == "__main__":
    main()
